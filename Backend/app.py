
import calendar
from io import BytesIO
import os
import re
import traceback
import unicodedata
import uuid
from datetime import date, datetime, timedelta
from calendar import monthrange
from typing import Dict, Optional
from flask import Flask, jsonify, request, render_template, redirect, send_file
from flask_cors import CORS
from azure.data.tables import TableServiceClient, UpdateMode
from azure.core.exceptions import ResourceNotFoundError
import openpyxl

app = Flask(__name__)
CORS(app)  # Importante para que Angular pueda consultar al Backend

TABLE_NAME = os.getenv('AZURE_TABLE_NAME', 'baterias')
STORAGE_CONNECTION_STRING = os.getenv('AZURE_STORAGE_CONNECTION_STRING', '').strip()
EXCEL_DATA_FILE = os.getenv(
    'EXCEL_DATA_FILE',
    os.path.join(os.path.dirname(__file__), 'data', 'plantilla_baterias.xlsx'),
)
BALANZAS_TABLE_NAME = os.getenv('AZURE_BALANZAS_TABLE_NAME', 'balanzas')
BALANZAS_EXCEL_DATA_FILE = os.getenv(
    'BALANZAS_EXCEL_DATA_FILE',
    os.path.join(os.path.dirname(__file__), 'data', 'Balanzas.xlsx'),
)
FRONTEND_BASE_URL = os.getenv('FRONTEND_BASE_URL', '').strip()
PARTITION_KEY = 'baterias'
BALANZAS_PARTITION_KEY = 'balanzas'
ROW_ID_COLUMN = 'rowId'

COLUMNAS = [
    'COD',
    'Negocio',
    'UPS Marca',
    'Modelo',
    'Capacidad',
    'Serial',
    'Inventario No',
    'FECHA DE INSTALACION',
    'REFERENCIA',
    'CANTIDAD',
    'FECHA DE VENCIMIENTO',
    'ESTADO',
    'DIAS VENCIDOS',
    'AÑOS/ MESES/ DIAS',
]

BALANZAS_COLUMNAS = [
    'COD',
    'Negocio',
    'Marca',
    'Modelo',
    'Ubicación',
    'Activo',
    'Serial',
    'FECHA CERTIFICACION - COMPRA',
    'NII',
    'FECHA DE VENCIMIENTO',
    'ESTADO',
    'DIAS VENCIDOS',
    'ACTAS',
    'OBSERVACIONES',
]


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize('NFD', value)
    normalized = ''.join(
        ch for ch in normalized if unicodedata.category(ch) != 'Mn'
    )
    normalized = normalized.lower()
    normalized = re.sub(r'[^a-z0-9]+', '_', normalized)
    return normalized.strip('_')


COLUMN_KEY_MAP = {col: normalize_key(col) for col in COLUMNAS}
BALANZAS_COLUMN_KEY_MAP = {col: normalize_key(col) for col in BALANZAS_COLUMNAS}


def get_table_client():
    if not STORAGE_CONNECTION_STRING:
        raise ValueError('AZURE_STORAGE_CONNECTION_STRING no esta configurado.')

    service = TableServiceClient.from_connection_string(STORAGE_CONNECTION_STRING)
    service.create_table_if_not_exists(TABLE_NAME)
    return service.get_table_client(TABLE_NAME)


def using_azure_table_storage() -> bool:
    return bool(STORAGE_CONNECTION_STRING)


def get_frontend_base_url() -> Optional[str]:
    if FRONTEND_BASE_URL:
        return FRONTEND_BASE_URL.rstrip('/')

    if not using_azure_table_storage():
        return 'http://172.19.72.16:4200'

    return None


def redirect_to_frontend_or_api(path: str = ''):
    frontend_url = get_frontend_base_url()
    if frontend_url:
        return redirect(f"{frontend_url}{path}", code=302)

    return jsonify({
        'ok': True,
        'message': 'Backend API activo. Esta URL ya no sirve la interfaz web.',
        'api': {
            'baterias': '/api/baterias',
            'sync': '/api/baterias/sync',
            'balanzas': '/api/balanzas',
            'balanzas_sync': '/api/balanzas/sync',
        },
    }), 200


def ensure_excel_file_exists() -> None:
    if not os.path.exists(EXCEL_DATA_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([*COLUMNAS, ROW_ID_COLUMN])
        workbook.save(EXCEL_DATA_FILE)


def normalize_header(value: Optional[str]) -> str:
    if value is None:
        return ''
    return str(value).strip()


def load_excel_records() -> list[dict[str, str]]:
    ensure_excel_file_exists()

    workbook = openpyxl.load_workbook(EXCEL_DATA_FILE)
    sheet = workbook.active
    headers = [normalize_header(cell.value) for cell in sheet[1]]

    if not any(headers):
        headers = [ROW_ID_COLUMN, *COLUMNAS]
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column_index).value = header

    row_id_index = next(
        (index for index, header in enumerate(headers) if header.lower() == ROW_ID_COLUMN.lower()),
        -1,
    )

    if row_id_index == -1:
        headers.append(ROW_ID_COLUMN)
        row_id_index = len(headers) - 1
        sheet.cell(row=1, column=len(headers)).value = ROW_ID_COLUMN

    rows: list[dict[str, str]] = []
    changed = False

    for row_index in range(2, sheet.max_row + 1):
        row_values: dict[str, str] = {}
        is_empty = True

        for column_index, header in enumerate(headers, start=1):
            cell_value = sheet.cell(row=row_index, column=column_index).value
            normalized = '' if cell_value is None else str(cell_value)
            row_values[header] = normalized
            if normalized.strip():
                is_empty = False

        if is_empty:
            continue

        if not row_values.get(ROW_ID_COLUMN):
            row_values[ROW_ID_COLUMN] = uuid.uuid4().hex
            sheet.cell(row=row_index, column=row_id_index + 1).value = row_values[ROW_ID_COLUMN]
            changed = True

        rows.append(row_values)

    if changed:
        workbook.save(EXCEL_DATA_FILE)

    return rows


def save_excel_records(records: list[dict[str, str]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = [ROW_ID_COLUMN, *COLUMNAS]
    sheet.append(headers)

    for record in records:
        sheet.append([record.get(header, '') for header in headers])

    try:
        workbook.save(EXCEL_DATA_FILE)
    except PermissionError as error:
        raise IOError(
            f"No se pudo guardar el archivo Excel '{EXCEL_DATA_FILE}'. Ciérrelo si está abierto en otra aplicación."
        ) from error


def build_excel_record(payload: dict, row_key: str) -> dict[str, str]:
    record = {ROW_ID_COLUMN: row_key}
    for col in COLUMNAS:
        record[col] = normalize_payload_value(get_payload_value(payload, col))
    return compute_excel_fields(record)


def find_excel_record_index(records: list[dict[str, str]], row_key: str) -> int:
    for index, record in enumerate(records):
        if record.get(ROW_ID_COLUMN) == row_key:
            return index
    return -1


def normalize_payload_value(value):
    if value is None:
        return ""
    return str(value)


def get_payload_value(payload: dict, expected_key: str):
    if expected_key in payload:
        return payload[expected_key]

    normalized_key = normalize_key(expected_key)
    for key in payload:
        if key == expected_key:
            return payload[key]
        if normalize_key(str(key)) == normalized_key:
            return payload[key]
        if str(key).strip().lower() == expected_key.strip().lower():
            return payload[key]

    return ""


def parse_date_string(value: str) -> Optional[date]:
    if not value:
        return None

    value_str = str(value).strip()
    if not value_str:
        return None

    for fmt in (
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%Y/%m/%d',
        '%d-%m-%Y',
        '%m/%d/%Y',
        '%Y.%m.%d',
    ):
        try:
            return datetime.strptime(value_str, fmt).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(value_str).date()
    except ValueError:
        return None


def add_months(base_date: date, months: int) -> date:
    month = base_date.month - 1 + months
    year = base_date.year + month // 12
    month = month % 12 + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def diff_in_years_months_days(start: date, end: date) -> dict[str, int]:
    years = end.year - start.year
    months = end.month - start.month
    days = end.day - start.day

    if days < 0:
        previous_month = date(end.year, end.month, 1) - timedelta(days=1)
        days += previous_month.day
        months -= 1

    if months < 0:
        months += 12
        years -= 1

    return {'years': years, 'months': months, 'days': days}


def compute_excel_fields(record: dict[str, str]) -> dict[str, str]:
    due_date = parse_date_string(record.get('FECHA DE VENCIMIENTO', ''))
    if not due_date:
        install_date = parse_date_string(record.get('FECHA DE INSTALACION', ''))
        if install_date:
            due_date = add_months(install_date, 36)
            record['FECHA DE VENCIMIENTO'] = due_date.strftime('%Y-%m-%d')

    if due_date is None:
        return record

    today = date.today()
    diff_days = (due_date - today).days
    record['DIAS VENCIDOS'] = str(diff_days)

    if diff_days < 0:
        record['ESTADO'] = 'Vencido'
    elif diff_days <= 30:
        record['ESTADO'] = 'Por vencer'
    else:
        record['ESTADO'] = 'Vigente'

    start, end = (due_date, today) if due_date < today else (today, due_date)
    delta = diff_in_years_months_days(start, end)
    record['AÑOS/ MESES/ DIAS'] = f"{delta['years']} AÑOS {delta['months']} MESES {delta['days']} DIAS"
    return record


def build_entity(payload: dict, row_key: str) -> dict:
    entity = {'PartitionKey': PARTITION_KEY, 'RowKey': row_key}
    for col in COLUMNAS:
        column_key = COLUMN_KEY_MAP[col]
        entity[column_key] = normalize_payload_value(get_payload_value(payload, col))
    return entity


def record_from_entity(entity: dict) -> dict:
    record = {col: str(entity.get(COLUMN_KEY_MAP[col], "")) for col in COLUMNAS}
    record['rowId'] = entity.get('RowKey', '')
    return record



def parse_date(value: str) -> Optional[date]:
    if not value:
        return None

    clean = str(value).strip()
    iso_match = re.match(r'^(\d{4})-(\d{2})-(\d{2})', clean)
    if iso_match:
        year, month, day = map(int, iso_match.groups())
        try:
            return date(year, month, day)
        except ValueError:
            return None

    for fmt in ('%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(clean, fmt).date()
        except ValueError:
            continue
    return None


def add_months(base: date, months: int) -> date:
    month = base.month - 1 + months
    year = base.year + month // 12
    month = month % 12 + 1
    day = min(base.day, monthrange(year, month)[1])
    return date(year, month, day)


def compute_estado(diff_days: int) -> str:
    if diff_days < 0:
        return 'Vencido'
    if diff_days <= 30:
        return 'Por vencer'
    return 'Vigente'


def compute_tiempo_detalle(start: date, end: date) -> str:
    years = end.year - start.year
    months = end.month - start.month
    days = end.day - start.day

    if days < 0:
        prev_month = end.month - 1 or 12
        prev_year = end.year - 1 if end.month == 1 else end.year
        days += monthrange(prev_year, prev_month)[1]
        months -= 1

    if months < 0:
        months += 12
        years -= 1

    return f'{years} AÑOS {months} MESES {days} DIAS'


def build_computed_values(fecha_instalacion_raw: str) -> Dict[str, str]:
    fecha_instalacion = parse_date(fecha_instalacion_raw)
    if not fecha_instalacion:
        return {
            'FECHA DE VENCIMIENTO': '',
            'ESTADO': '',
            'DIAS VENCIDOS': '',
            'AÑOS/ MESES/ DIAS': '',
        }

    fecha_vencimiento = add_months(fecha_instalacion, 36)
    today = date.today()
    diff_days = (fecha_vencimiento - today).days
    start, end = (fecha_vencimiento, today) if fecha_vencimiento < today else (today, fecha_vencimiento)

    return {
        'FECHA DE VENCIMIENTO': fecha_vencimiento.isoformat(),
        'ESTADO': compute_estado(diff_days),
        'DIAS VENCIDOS': str(diff_days),
        'AÑOS/ MESES/ DIAS': compute_tiempo_detalle(start, end),
    }


def get_balanzas_table_client():
    if not STORAGE_CONNECTION_STRING:
        raise ValueError('AZURE_STORAGE_CONNECTION_STRING no esta configurado.')

    service = TableServiceClient.from_connection_string(STORAGE_CONNECTION_STRING)
    return service.get_table_client(BALANZAS_TABLE_NAME)


def format_excel_cell_value(value) -> str:
    if value is None:
        return ''

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return str(value).strip()


def ensure_balanzas_excel_file_exists() -> None:
    if not os.path.exists(BALANZAS_EXCEL_DATA_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append([*BALANZAS_COLUMNAS, ROW_ID_COLUMN])
        workbook.save(BALANZAS_EXCEL_DATA_FILE)


def find_balanzas_header_row(sheet) -> int:
    max_probe = min(sheet.max_row, 25)

    for row_index in range(1, max_probe + 1):
        row_values = [normalize_header(sheet.cell(row=row_index, column=col_index).value) for col_index in range(1, max(15, sheet.max_column) + 1)]
        normalized = {normalize_key(value) for value in row_values if value}
        if {'cod', 'negocio'}.issubset(normalized):
            return row_index

    return 1


def load_balanzas_records() -> list[dict[str, str]]:
    ensure_balanzas_excel_file_exists()

    workbook = openpyxl.load_workbook(BALANZAS_EXCEL_DATA_FILE)
    sheet = workbook.active

    header_row_index = find_balanzas_header_row(sheet)
    headers = [normalize_header(cell.value) for cell in sheet[header_row_index]]

    if not any(headers):
        headers = [*BALANZAS_COLUMNAS, ROW_ID_COLUMN]
        header_row_index = 1
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=header_row_index, column=column_index).value = header

    row_id_index = next(
        (index for index, header in enumerate(headers) if header.lower() == ROW_ID_COLUMN.lower()),
        -1,
    )

    if row_id_index == -1:
        headers.append(ROW_ID_COLUMN)
        row_id_index = len(headers) - 1
        sheet.cell(row=header_row_index, column=len(headers)).value = ROW_ID_COLUMN

    rows: list[dict[str, str]] = []
    changed = False

    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        row_values: dict[str, str] = {}
        is_empty = True

        for column_index, header in enumerate(headers, start=1):
            value = format_excel_cell_value(sheet.cell(row=row_index, column=column_index).value)
            row_values[header] = value
            if value.strip():
                is_empty = False

        if is_empty:
            continue

        if not row_values.get(ROW_ID_COLUMN):
            row_values[ROW_ID_COLUMN] = uuid.uuid4().hex
            sheet.cell(row=row_index, column=row_id_index + 1).value = row_values[ROW_ID_COLUMN]
            changed = True

        rows.append(row_values)

    if changed:
        workbook.save(BALANZAS_EXCEL_DATA_FILE)

    return rows


def save_balanzas_records(records: list[dict[str, str]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = [*BALANZAS_COLUMNAS, ROW_ID_COLUMN]
    sheet.append(headers)

    for record in records:
        sheet.append([record.get(header, '') for header in headers])

    try:
        workbook.save(BALANZAS_EXCEL_DATA_FILE)
    except PermissionError as error:
        raise IOError(
            f"No se pudo guardar el archivo Excel '{BALANZAS_EXCEL_DATA_FILE}'. Ciérrelo si está abierto en otra aplicación."
        ) from error


def compute_balanzas_fields(record: dict[str, str]) -> dict[str, str]:
    due_date = parse_date_string(record.get('FECHA DE VENCIMIENTO', ''))
    if not due_date:
        cert_date = parse_date_string(record.get('FECHA CERTIFICACION - COMPRA', ''))
        if cert_date:
            due_date = add_months(cert_date, 24)
            record['FECHA DE VENCIMIENTO'] = due_date.strftime('%Y-%m-%d')

    if due_date is None:
        return record

    today = date.today()
    diff_days = (due_date - today).days
    record['DIAS VENCIDOS'] = str(diff_days)
    record['ESTADO'] = compute_estado(diff_days)
    return record


def build_balanzas_record(payload: dict, row_key: str) -> dict[str, str]:
    record = {ROW_ID_COLUMN: row_key}
    for col in BALANZAS_COLUMNAS:
        record[col] = normalize_payload_value(get_payload_value(payload, col))
    return compute_balanzas_fields(record)


def find_balanzas_record_index(records: list[dict[str, str]], row_key: str) -> int:
    for index, record in enumerate(records):
        if record.get(ROW_ID_COLUMN) == row_key:
            return index
    return -1


def build_balanzas_entity(payload: dict, row_key: str) -> dict:
    entity = {'PartitionKey': BALANZAS_PARTITION_KEY, 'RowKey': row_key}
    for col in BALANZAS_COLUMNAS:
        column_key = BALANZAS_COLUMN_KEY_MAP[col]
        entity[column_key] = normalize_payload_value(get_payload_value(payload, col))
    return entity


def balanza_record_from_entity(entity: dict) -> dict:
    record = {col: str(entity.get(BALANZAS_COLUMN_KEY_MAP[col], '')) for col in BALANZAS_COLUMNAS}
    record['rowId'] = entity.get('RowKey', '')
    return record


def build_balanzas_computed_values(fecha_certificacion_raw: str) -> Dict[str, str]:
    fecha_certificacion = parse_date(fecha_certificacion_raw)
    if not fecha_certificacion:
        return {
            'FECHA DE VENCIMIENTO': '',
            'ESTADO': '',
            'DIAS VENCIDOS': '',
        }

    fecha_vencimiento = add_months(fecha_certificacion, 24)
    today = date.today()
    diff_days = (fecha_vencimiento - today).days

    return {
        'FECHA DE VENCIMIENTO': fecha_vencimiento.isoformat(),
        'ESTADO': compute_estado(diff_days),
        'DIAS VENCIDOS': str(diff_days),
    }


@app.route('/')
def index():
    return redirect_to_frontend_or_api()

@app.route('/gestion_baterias')
def gestion_baterias():
    return redirect_to_frontend_or_api('/modulos/baterias')


@app.route('/gestion_balanzas')
def gestion_balanzas():
    return redirect_to_frontend_or_api('/modulos/balanzas')

@app.route('/agregar')
def agregar():
    return redirect_to_frontend_or_api('/modulos/baterias/agregar')


@app.route('/editar')
def editar():
    return redirect_to_frontend_or_api('/modulos/baterias/editar')

@app.route('/eliminar')
def eliminar():
    return redirect_to_frontend_or_api('/modulos/baterias/eliminar')


@app.route('/balanzas/agregar')
def agregar_balanzas():
    return redirect_to_frontend_or_api('/modulos/balanzas/agregar')


@app.route('/balanzas/editar')
def editar_balanzas():
    return redirect_to_frontend_or_api('/modulos/balanzas/editar')


@app.route('/balanzas/eliminar')
def eliminar_balanzas():
    return redirect_to_frontend_or_api('/modulos/balanzas/eliminar')
    
@app.route('/api/baterias', methods=['GET'])
def listar_baterias():
    try:
        if using_azure_table_storage():
            table_client = get_table_client()
            include_all = request.args.get('all', '').strip().lower() in {'1', 'true', 'yes'}
            if include_all:
                entities = table_client.list_entities()
            else:
                entities = table_client.query_entities(
                    f"PartitionKey eq '{PARTITION_KEY}'"
                )
            data = [record_from_entity(entity) for entity in entities]
        else:
            data = load_excel_records()
        return jsonify(data), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al consultar los registros: {str(e)}"}), 500


@app.route('/api/baterias/sync', methods=['POST'])

def sincronizar_baterias():
    try:
        if using_azure_table_storage():
            table_client = get_table_client()
            entities = list(table_client.query_entities(
                f"PartitionKey eq '{PARTITION_KEY}'"
            ))

            updated_count = 0
            for entity in entities:
                fecha_instalacion = str(entity.get(COLUMN_KEY_MAP['FECHA DE INSTALACION'], '')).strip()
                computed = build_computed_values(fecha_instalacion)

                patch = {
                    'PartitionKey': entity['PartitionKey'],
                    'RowKey': entity['RowKey'],
                }
                changed = False
                for col_name, value in computed.items():
                    key = COLUMN_KEY_MAP[col_name]
                    current_value = str(entity.get(key, ''))
                    if current_value != value:
                        patch[key] = value
                        changed = True

                if changed:
                    table_client.update_entity(patch, mode=UpdateMode.MERGE)
                    updated_count += 1

            return jsonify({
                'ok': True,
                'updated': updated_count,
                'total': len(entities),
                'source': 'azure-table',
            }), 200

        records = load_excel_records()
        updated_count = 0
        synced_records: list[dict[str, str]] = []

        for record in records:
            before = {
                'FECHA DE VENCIMIENTO': str(record.get('FECHA DE VENCIMIENTO', '')),
                'ESTADO': str(record.get('ESTADO', '')),
                'DIAS VENCIDOS': str(record.get('DIAS VENCIDOS', '')),
                'AÑOS/ MESES/ DIAS': str(record.get('AÑOS/ MESES/ DIAS', '')),
            }
            updated = compute_excel_fields(dict(record))
            after = {
                'FECHA DE VENCIMIENTO': str(updated.get('FECHA DE VENCIMIENTO', '')),
                'ESTADO': str(updated.get('ESTADO', '')),
                'DIAS VENCIDOS': str(updated.get('DIAS VENCIDOS', '')),
                'AÑOS/ MESES/ DIAS': str(updated.get('AÑOS/ MESES/ DIAS', '')),
            }
            if before != after:
                updated_count += 1
            synced_records.append(updated)

        save_excel_records(synced_records)
        return jsonify({
            'ok': True,
            'updated': updated_count,
            'total': len(records),
            'source': 'excel',
        }), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al sincronizar los registros: {str(e)}"}), 500


@app.route('/api/baterias/export', methods=['GET'])
def exportar_baterias_excel():
    try:
        filtro_estado_raw = (request.args.get('estado') or 'all').strip()

        def normalize_estado(value: str) -> str:
            normalized = unicodedata.normalize('NFD', value)
            normalized = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
            normalized = normalized.strip().lower()
            return re.sub(r'\s+', ' ', normalized)

        def matches_estado_filter(record: dict) -> bool:
            normalized_filter = normalize_estado(filtro_estado_raw)
            if normalized_filter in {'', 'all', 'todo', 'todos'}:
                return True

            estado = str(record.get('ESTADO', '')).strip()
            normalized_estado = normalize_estado(estado)

            aliases = {
                'vigente': {'vigente'},
                'por vencer': {'por vencer', 'porvencer'},
                'vencido': {'vencido', 'vencidos'},
            }

            for canonical, options in aliases.items():
                if normalized_filter in options:
                    return normalized_estado == canonical

            return normalized_estado == normalized_filter

        def build_export_file_name() -> str:
            normalized_filter = normalize_estado(filtro_estado_raw)
            suffix_map = {
                'all': 'todo',
                'todo': 'todo',
                'todos': 'todo',
                'vigente': 'vigente',
                'por vencer': 'por_vencer',
                'porvencer': 'por_vencer',
                'vencido': 'vencido',
                'vencidos': 'vencido',
            }
            suffix = suffix_map.get(normalized_filter, 'filtro')
            return f'baterias_export_{suffix}.xlsx'

        if using_azure_table_storage():
            table_client = get_table_client()
            entities = list(table_client.query_entities(
                f"PartitionKey eq '{PARTITION_KEY}'"
            ))
            records = [record_from_entity(entity) for entity in entities]
            records = [record for record in records if matches_estado_filter(record)]

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = [*COLUMNAS, ROW_ID_COLUMN]
            sheet.append(headers)

            for record in records:
                sheet.append([
                    record.get('COD', ''),
                    record.get('Negocio', ''),
                    record.get('UPS Marca', ''),
                    record.get('Modelo', ''),
                    record.get('Capacidad', ''),
                    record.get('Serial', ''),
                    record.get('Inventario No', ''),
                    record.get('FECHA DE INSTALACION', ''),
                    record.get('REFERENCIA', ''),
                    record.get('CANTIDAD', ''),
                    record.get('FECHA DE VENCIMIENTO', ''),
                    record.get('ESTADO', ''),
                    record.get('DIAS VENCIDOS', ''),
                    record.get('AÑOS/ MESES/ DIAS', ''),
                    record.get('rowId', ''),
                ])

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return send_file(
                output,
                as_attachment=True,
                download_name=build_export_file_name(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )

        records = load_excel_records()
        updated_records = [compute_excel_fields(dict(record)) for record in records]
        save_excel_records(updated_records)
        filtered_records = [record for record in updated_records if matches_estado_filter(record)]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = [*COLUMNAS, ROW_ID_COLUMN]
        sheet.append(headers)

        for record in filtered_records:
            sheet.append([
                record.get('COD', ''),
                record.get('Negocio', ''),
                record.get('UPS Marca', ''),
                record.get('Modelo', ''),
                record.get('Capacidad', ''),
                record.get('Serial', ''),
                record.get('Inventario No', ''),
                record.get('FECHA DE INSTALACION', ''),
                record.get('REFERENCIA', ''),
                record.get('CANTIDAD', ''),
                record.get('FECHA DE VENCIMIENTO', ''),
                record.get('ESTADO', ''),
                record.get('DIAS VENCIDOS', ''),
                record.get('AÑOS/ MESES/ DIAS', ''),
                record.get('rowId', ''),
            ])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=build_export_file_name(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al exportar Excel: {str(e)}"}), 500



@app.route('/api/baterias', methods=['POST'])
def crear_bateria():
    try:
        payload = request.get_json(silent=True) or {}
        if not payload:
            return jsonify({"error": "Carga JSON inválida o cuerpo vacío."}), 400

        row_key = uuid.uuid4().hex

        if using_azure_table_storage():
            table_client = get_table_client()
            entity = build_entity(payload, row_key)
            table_client.create_entity(entity)
        else:
            record = build_excel_record(payload, row_key)
            records = load_excel_records()
            records.append(record)
            save_excel_records(records)

        return jsonify({"ok": True, "rowId": row_key}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al guardar el registro: {str(e)}"}), 500


@app.route('/api/baterias/<row_id>', methods=['PUT'])
def actualizar_bateria(row_id: str):
    try:
        payload = request.get_json(silent=True) or {}

        if using_azure_table_storage():
            table_client = get_table_client()
            try:
                table_client.get_entity(PARTITION_KEY, row_id)
            except ResourceNotFoundError:
                return jsonify({"error": "Registro no encontrado"}), 404

            entity = build_entity(payload, row_id)
            table_client.update_entity(entity, mode=UpdateMode.REPLACE)
        else:
            records = load_excel_records()
            index = find_excel_record_index(records, row_id)
            if index == -1:
                return jsonify({"error": "Registro no encontrado"}), 404
            records[index] = build_excel_record(payload, row_id)
            save_excel_records(records)

        return jsonify({"ok": True, "mensaje": "Registro actualizado"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al actualizar el registro: {str(e)}"}), 500


@app.route('/api/baterias/<row_id>', methods=['DELETE'])
def eliminar_bateria(row_id: str):
    try:
        if using_azure_table_storage():
            table_client = get_table_client()

            try:
                table_client.get_entity(PARTITION_KEY, row_id)
            except ResourceNotFoundError:
                return jsonify({"error": "Registro no encontrado"}), 404

            table_client.delete_entity(PARTITION_KEY, row_id)
        else:
            records = load_excel_records()
            index = find_excel_record_index(records, row_id)
            if index == -1:
                return jsonify({"error": "Registro no encontrado"}), 404
            records.pop(index)
            save_excel_records(records)

        return jsonify({"ok": True, "mensaje": "Registro eliminado"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al eliminar el registro: {str(e)}"}), 500


@app.route('/api/balanzas', methods=['GET'])
def listar_balanzas():
    try:
        if using_azure_table_storage():
            try:
                table_client = get_balanzas_table_client()
                include_all = request.args.get('all', '').strip().lower() in {'1', 'true', 'yes'}
                if include_all:
                    entities = table_client.list_entities()
                else:
                    entities = table_client.query_entities(
                        f"PartitionKey eq '{BALANZAS_PARTITION_KEY}'"
                    )
                data = [balanza_record_from_entity(entity) for entity in entities]
            except Exception:
                traceback.print_exc()
                data = load_balanzas_records()
        else:
            data = load_balanzas_records()
        return jsonify(data), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al consultar los registros: {str(e)}"}), 500


@app.route('/api/balanzas/sync', methods=['POST'])
def sincronizar_balanzas():
    try:
        if using_azure_table_storage():
            try:
                table_client = get_balanzas_table_client()
                entities = list(table_client.query_entities(
                    f"PartitionKey eq '{BALANZAS_PARTITION_KEY}'"
                ))

                updated_count = 0
                for entity in entities:
                    fecha_certificacion = str(entity.get(BALANZAS_COLUMN_KEY_MAP['FECHA CERTIFICACION - COMPRA'], '')).strip()
                    computed = build_balanzas_computed_values(fecha_certificacion)

                    patch = {
                        'PartitionKey': entity['PartitionKey'],
                        'RowKey': entity['RowKey'],
                    }
                    changed = False
                    for col_name, value in computed.items():
                        key = BALANZAS_COLUMN_KEY_MAP[col_name]
                        current_value = str(entity.get(key, ''))
                        if current_value != value:
                            patch[key] = value
                            changed = True

                    if changed:
                        table_client.update_entity(patch, mode=UpdateMode.MERGE)
                        updated_count += 1

                return jsonify({
                    'ok': True,
                    'updated': updated_count,
                    'total': len(entities),
                    'source': 'azure-table',
                }), 200
            except Exception:
                traceback.print_exc()

        records = load_balanzas_records()
        updated_count = 0
        synced_records: list[dict[str, str]] = []

        for record in records:
            before = {
                'FECHA DE VENCIMIENTO': str(record.get('FECHA DE VENCIMIENTO', '')),
                'ESTADO': str(record.get('ESTADO', '')),
                'DIAS VENCIDOS': str(record.get('DIAS VENCIDOS', '')),
            }
            updated = compute_balanzas_fields(dict(record))
            after = {
                'FECHA DE VENCIMIENTO': str(updated.get('FECHA DE VENCIMIENTO', '')),
                'ESTADO': str(updated.get('ESTADO', '')),
                'DIAS VENCIDOS': str(updated.get('DIAS VENCIDOS', '')),
            }
            if before != after:
                updated_count += 1
            synced_records.append(updated)

        save_balanzas_records(synced_records)
        return jsonify({
            'ok': True,
            'updated': updated_count,
            'total': len(records),
            'source': 'excel',
        }), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al sincronizar los registros: {str(e)}"}), 500


@app.route('/api/balanzas/export', methods=['GET'])
def exportar_balanzas_excel():
    try:
        filtro_estado_raw = (request.args.get('estado') or 'all').strip()

        def normalize_estado(value: str) -> str:
            normalized = unicodedata.normalize('NFD', value)
            normalized = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
            normalized = normalized.strip().lower()
            return re.sub(r'\s+', ' ', normalized)

        def matches_estado_filter(record: dict) -> bool:
            normalized_filter = normalize_estado(filtro_estado_raw)
            if normalized_filter in {'', 'all', 'todo', 'todos'}:
                return True

            estado = str(record.get('ESTADO', '')).strip()
            normalized_estado = normalize_estado(estado)

            aliases = {
                'vigente': {'vigente'},
                'por vencer': {'por vencer', 'porvencer'},
                'vencido': {'vencido', 'vencidos'},
            }

            for canonical, options in aliases.items():
                if normalized_filter in options:
                    return normalized_estado == canonical

            return normalized_estado == normalized_filter

        def build_export_file_name() -> str:
            normalized_filter = normalize_estado(filtro_estado_raw)
            suffix_map = {
                'all': 'todo',
                'todo': 'todo',
                'todos': 'todo',
                'vigente': 'vigente',
                'por vencer': 'por_vencer',
                'porvencer': 'por_vencer',
                'vencido': 'vencido',
                'vencidos': 'vencido',
            }
            suffix = suffix_map.get(normalized_filter, 'filtro')
            return f'balanzas_export_{suffix}.xlsx'

        if using_azure_table_storage():
            try:
                table_client = get_balanzas_table_client()
                entities = list(table_client.query_entities(
                    f"PartitionKey eq '{BALANZAS_PARTITION_KEY}'"
                ))
                records = [balanza_record_from_entity(entity) for entity in entities]
                records = [record for record in records if matches_estado_filter(record)]

                workbook = openpyxl.Workbook()
                sheet = workbook.active
                headers = [*BALANZAS_COLUMNAS, ROW_ID_COLUMN]
                sheet.append(headers)

                for record in records:
                    sheet.append([record.get(header, '') for header in headers])

                output = BytesIO()
                workbook.save(output)
                output.seek(0)
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=build_export_file_name(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
            except Exception:
                traceback.print_exc()

        records = load_balanzas_records()
        updated_records = [compute_balanzas_fields(dict(record)) for record in records]
        save_balanzas_records(updated_records)
        filtered_records = [record for record in updated_records if matches_estado_filter(record)]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = [*BALANZAS_COLUMNAS, ROW_ID_COLUMN]
        sheet.append(headers)

        for record in filtered_records:
            sheet.append([record.get(header, '') for header in headers])

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name=build_export_file_name(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al exportar Excel: {str(e)}"}), 500


@app.route('/api/balanzas', methods=['POST'])
def crear_balanza():
    try:
        payload = request.get_json(silent=True) or {}
        if not payload:
            return jsonify({"error": "Carga JSON inválida o cuerpo vacío."}), 400

        row_key = uuid.uuid4().hex

        if using_azure_table_storage():
            try:
                table_client = get_balanzas_table_client()
                entity = build_balanzas_entity(payload, row_key)
                table_client.create_entity(entity)
            except Exception:
                traceback.print_exc()
                record = build_balanzas_record(payload, row_key)
                records = load_balanzas_records()
                records.append(record)
                save_balanzas_records(records)
        else:
            record = build_balanzas_record(payload, row_key)
            records = load_balanzas_records()
            records.append(record)
            save_balanzas_records(records)

        return jsonify({"ok": True, "rowId": row_key}), 201
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al guardar el registro: {str(e)}"}), 500


@app.route('/api/balanzas/<row_id>', methods=['PUT'])
def actualizar_balanza(row_id: str):
    try:
        payload = request.get_json(silent=True) or {}

        if using_azure_table_storage():
            try:
                table_client = get_balanzas_table_client()
                try:
                    table_client.get_entity(BALANZAS_PARTITION_KEY, row_id)
                except ResourceNotFoundError:
                    return jsonify({"error": "Registro no encontrado"}), 404

                entity = build_balanzas_entity(payload, row_id)
                table_client.update_entity(entity, mode=UpdateMode.REPLACE)
            except Exception:
                traceback.print_exc()
                records = load_balanzas_records()
                index = find_balanzas_record_index(records, row_id)
                if index == -1:
                    return jsonify({"error": "Registro no encontrado"}), 404
                records[index] = build_balanzas_record(payload, row_id)
                save_balanzas_records(records)
        else:
            records = load_balanzas_records()
            index = find_balanzas_record_index(records, row_id)
            if index == -1:
                return jsonify({"error": "Registro no encontrado"}), 404
            records[index] = build_balanzas_record(payload, row_id)
            save_balanzas_records(records)

        return jsonify({"ok": True, "mensaje": "Registro actualizado"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al actualizar el registro: {str(e)}"}), 500


@app.route('/api/balanzas/<row_id>', methods=['DELETE'])
def eliminar_balanza(row_id: str):
    try:
        if using_azure_table_storage():
            try:
                table_client = get_balanzas_table_client()

                try:
                    table_client.get_entity(BALANZAS_PARTITION_KEY, row_id)
                except ResourceNotFoundError:
                    return jsonify({"error": "Registro no encontrado"}), 404

                table_client.delete_entity(BALANZAS_PARTITION_KEY, row_id)
            except Exception:
                traceback.print_exc()
                records = load_balanzas_records()
                index = find_balanzas_record_index(records, row_id)
                if index == -1:
                    return jsonify({"error": "Registro no encontrado"}), 404
                records.pop(index)
                save_balanzas_records(records)
        else:
            records = load_balanzas_records()
            index = find_balanzas_record_index(records, row_id)
            if index == -1:
                return jsonify({"error": "Registro no encontrado"}), 404
            records.pop(index)
            save_balanzas_records(records)

        return jsonify({"ok": True, "mensaje": "Registro eliminado"}), 200
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Error al eliminar el registro: {str(e)}"}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.getenv('PORT', '8500')))