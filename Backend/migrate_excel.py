#!/usr/bin/env python3
"""
Script para migrar datos de plantilla_baterias.xlsx a Azure Table Storage.
Uso: python migrate_excel.py
"""
import os
import sys
import uuid
import unicodedata
import re
import openpyxl
from azure.data.tables import TableServiceClient


# Configuración idéntica a app.py
TABLE_NAME = os.getenv('AZURE_TABLE_NAME', 'baterias')
STORAGE_CONNECTION_STRING = os.getenv('AZURE_STORAGE_CONNECTION_STRING', '')
PARTITION_KEY = 'baterias'

COLUMNAS = [
    'COD',
    'Negocio',
    'UPS Marca',
    'Modelo',
    'Capacidad',
    'Serial',
    'Inventario No',
    'FECHA DE INSTALACION',
    'REFERENCIA',
    'CANTIDAD',
    'FECHA DE VENCIMIENTO',
    'ESTADO',
    'DIAS VENCIDOS',
    'AÑOS/ MESES/ DIAS',
]


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize('NFD', value)
    normalized = ''.join(
        ch for ch in normalized if unicodedata.category(ch) != 'Mn'
    )
    normalized = normalized.lower()
    normalized = re.sub(r'[^a-z0-9]+', '_', normalized)
    return normalized.strip('_')


COLUMN_KEY_MAP = {col: normalize_key(col) for col in COLUMNAS}


def normalize_payload_value(value):
    if value is None:
        return ""
    return str(value)


def build_entity(payload: dict, row_key: str) -> dict:
    entity = {'PartitionKey': PARTITION_KEY, 'RowKey': row_key}
    for col in COLUMNAS:
        column_key = COLUMN_KEY_MAP[col]
        entity[column_key] = normalize_payload_value(payload.get(col, ""))
    return entity


def get_table_client():
    if not STORAGE_CONNECTION_STRING:
        raise ValueError('AZURE_STORAGE_CONNECTION_STRING no está configurado.')
    
    service = TableServiceClient.from_connection_string(STORAGE_CONNECTION_STRING)
    service.create_table_if_not_exists(TABLE_NAME)
    return service.get_table_client(TABLE_NAME)


def migrate_excel_to_azure(excel_path: str):
    """Lee Excel y sube datos a Azure Table Storage."""
    
    # Verificar configuración
    if not STORAGE_CONNECTION_STRING:
        print("❌ Error: Variable AZURE_STORAGE_CONNECTION_STRING no configurada.")
        print("   Asígnalas en tu terminal antes de ejecutar este script:")
        print("   $env:AZURE_STORAGE_CONNECTION_STRING='tu_connection_string'")
        sys.exit(1)
    
    # Cargar Excel
    print(f"📂 Leyendo Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    print(f"📋 Hoja: {ws.title}, Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    # Conectar a Azure
    print("🔗 Conectando a Azure Table Storage...")
    try:
        table_client = get_table_client()
        print(f"✅ Conexión OK, tabla: {TABLE_NAME}")
    except Exception as e:
        print(f"❌ Error de conexión: {e}")
        sys.exit(1)
    
    # Migrar filas (saltando header en fila 1)
    inserted = 0
    errors = 0
    
    for row_idx in range(2, ws.max_row + 1):
        try:
            # Leer fila
            row_data = {}
            for col_idx, col_name in enumerate(COLUMNAS, 1):
                cell_value = ws.cell(row_idx, col_idx).value
                row_data[col_name] = cell_value
            
            # Crear entity
            row_key = uuid.uuid4().hex
            entity = build_entity(row_data, row_key)
            
            # Insertar
            table_client.create_entity(entity)
            inserted += 1
            
            if inserted % 50 == 0:
                print(f"  ✓ Insertados: {inserted}/{ws.max_row - 1}")
        
        except Exception as e:
            errors += 1
            print(f"  ⚠️  Error en fila {row_idx}: {e}")
    
    print(f"\n📊 Resultado:")
    print(f"   ✅ Insertados: {inserted}")
    print(f"   ❌ Errores: {errors}")
    print(f"   Total: {inserted + errors}")
    
    if errors == 0:
        print("\n🎉 Migración completada exitosamente!")
    else:
        print(f"\n⚠️  Migración parcial. Revisa los {errors} errores arriba.")


if __name__ == '__main__':
    excel_file = os.path.join(os.path.dirname(__file__), 'data', 'plantilla_baterias.xlsx')
    
    if not os.path.exists(excel_file):
        print(f"❌ Archivo no encontrado: {excel_file}")
        sys.exit(1)
    
    migrate_excel_to_azure(excel_file)
